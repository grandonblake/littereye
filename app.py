from PySide6.QtCore import (QDate, QObject, Qt)
from PySide6.QtGui import (QImage, QPixmap)
from PySide6.QtWidgets import (QMainWindow, QDialog, QMessageBox, QFileDialog)

from PySide6 import QtWidgets

from PySide6.QtCore import Signal, QThread
from ultralytics import YOLO
import os, sys, datetime, time, cv2

from ui import Ui_MainWindow, Ui_settingsDialog, Ui_setAlertDialog
from database import Database

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from  matplotlib import rcParams

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

class VideoWorker(QObject):
    imageUpdated = Signal(QImage)
    errorOccurred = Signal(str)
    objectListUpdated = Signal(list)
    def __init__(self, capture, model_index, conf, iou, models, maxID, is_recording=False):
        super().__init__()
        self.capture = capture
        # self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        # self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)

        self.model_index = model_index
        self.conf = conf
        self.iou = iou
        self.models = [YOLO(model) for model in models]
        self.is_recording = is_recording # Flag to track recording state
        self.recording_out = None  # Video writer object
        self.maxID = maxID
        self.fps = 0.0
        self.recording_out = None

        self.names = {}

    def run(self):
        global start_time, frames

        frames = 0  # Initialize frame count
        start_time = time.time()  # Initialize start time 

        while True:
            ret, frame = self.capture.read()
            if ret:
                # Run YOLOv8 inference on the frame
                results = self.models[self.model_index].track(frame, stream=True, persist=True, conf=self.conf, iou=self.iou)
                detected_objects = {}

                for result in results:
                    # Visualize the results on the frame
                    frame = result.plot()
                    names = result.names 
                    classes = result.boxes.cls 
                    ids = result.boxes.id
                    conf = result.boxes.conf

                    self.names = result.names

                    # Check if ids and classes are not None
                    if ids is not None and classes is not None:
                        # Convert the tensors to lists
                        ids_list = ids.tolist()
                        classes_list = classes.tolist()
                        conf_list = conf.tolist()

                        # Iterate over both lists simultaneously
                        for i in range(len(ids_list)):
                            object_id = int(ids_list[i])  # Convert the ID to an integer
                            class_index = int(classes_list[i])  # Convert the class index to an integer
                            object_conf = conf_list[i]  # Get the confidence score as is
                            object_name = names[class_index]  # Get the corresponding name for the class index
                            detected_objects[object_id] = (object_name, object_conf)  # Assign the name and conf to the ID in the dictionary

                            datetimeNow = datetime.datetime.now()  # Get the current time
                            formattedDateTimeNow = datetimeNow.strftime("%m-%d-%Y %H-%M-%S")  # date and time in 24H format
                            
                            # {0: 'Cardboard', 1: 'Drink carton', 2: 'Face mask', 3: 'Glass bottle', 4: 'Metal can', 
                            # 5: 'Paper bag', 6: 'Paper container', 7: 'Paper cup', 8: 'Plastic bag', 9: 'Plastic bottle', 
                            # 10: 'Plastic cup', 11: 'Plastic wrapper', 12: 'Rags', 13: 'Styrofoam'}

                            if(class_index == 2 or class_index == 12 or class_index == 13): #not recyclable
                                recyclableBool = 0
                            else: #recyclable
                                recyclableBool = 1

                            maxID = self.maxID + object_id

                            with Database("insertObject") as database:
                                database.insertObject(objectID=maxID, className=object_name, confidenceLevel=object_conf, recyclableBool=recyclableBool, dateTime=formattedDateTimeNow)

                                # database.selectAllObjects()

                # Emit the detected objects list, even if it's empty
                self.objectListUpdated.emit(list(detected_objects.values()))

                # Calculate FPS
                fps = self.calculate_fps() 

                # Draw FPS text on the frame
                font = cv2.FONT_HERSHEY_SIMPLEX
                font_scale = 0.5
                font_thickness = 1
                text_color = (0, 0, 255)
                text_size, _ = cv2.getTextSize("FPS: {:.2f}".format(fps), font, font_scale, font_thickness)
                text_origin = (frame.shape[1] - text_size[0] - 10, text_size[1] + 30) 

                cv2.putText(frame, "FPS: {:.2f}".format(fps), text_origin, font, font_scale, text_color, font_thickness)

                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                image = QImage(frame, frame.shape[1], frame.shape[0], QImage.Format_RGB888)
                self.imageUpdated.emit(image.copy())

                if self.is_recording:
                    font = cv2.FONT_HERSHEY_SIMPLEX
                    font_scale = 0.5
                    font_thickness = 1
                    text_color = (255, 0, 0)
                    text = "RECORDING"
                    text_size, _ = cv2.getTextSize(text, font, font_scale, font_thickness)
                    text_origin = (10, text_size[1] + 10)  # Top-left corner with padding

                    cv2.putText(frame, text, text_origin, font, font_scale, text_color, font_thickness)

                    if self.recording_out is None:  # Initialize the video writer
                        now = datetime.datetime.now()  # Get the current time
                        filename = now.strftime("%m-%d-%Y %H-%M-%S.mp4")  # Format the filename
                        filepath = os.path.join('recordings', filename)  # Add the 'recordings' folder to the path
                        fourcc = cv2.VideoWriter_fourcc(*'mp4v')  
                        self.recording_out = cv2.VideoWriter(filepath, fourcc, fps, (frame.shape[1], frame.shape[0])) 

                    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # Convert the frame to RGB color space
                    self.recording_out.write(frame_rgb)  # Write the RGB frame to the video file

            else:
                self.errorOccurred.emit("Error reading from camera")
                break

            self.imageUpdated.emit(image.copy())

    def calculate_fps(self):
        global start_time, frames 
        frames += 1
        elapsed_time = time.time() - start_time
        if elapsed_time > 1:  # Update FPS every second
            self.fps = frames / elapsed_time
            frames = 0
            start_time = time.time()
        return self.fps
    
    def start_recording(self):
        self.is_recording = True
        
    def stop_recording(self):
        self.is_recording = False
        if self.recording_out is not None:
            self.recording_out.release()
            self.recording_out = None

    def open_alert_dialog(self):
        setAlert = setAlertDialog(self.parent(), self.names)
        setAlert.exec()

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.model_index = 0
        self.conf = 0.40
        self.iou = 0.50 
        self.models = self.getAvailableModels()  # list of models
        self.availableCameras = self.getAvailableCameras() #list of cameras
        self.maxID = self.maxIDFromDatabase()
        self.alert_dialog_shown = False
        
        self.StartRecordButton.clicked.connect(self.toggle_recording)  # Connect the button
        
        # Add available cameras to the CameraSourceComboBox
        for item in self.availableCameras:
            if isinstance(item, int):
                self.CameraSourceComboBox.addItem(f"Camera {item+1}", item)
            elif isinstance(item, str):
                self.CameraSourceComboBox.addItem(item, item)

        # Connect the currentIndexChanged signal to the runInference method
        self.CameraSourceComboBox.currentIndexChanged.connect(self.runInference)

        self.advancedSettingsButton.clicked.connect(self.onAdvancedSettingsClicked)

        self.SetAlertButton.clicked.connect(self.onSetAlertClicked)

        self.filterButton.clicked.connect(self.clickedFilterButton)

        self.updateDateEdits()

        self.fromDateEdit.dateChanged.connect(self.updateToDateMinimum)

        self.dailyRadioButton.toggled.connect(self.radioButtonToggled)
        self.hourlyRadioButton.toggled.connect(self.radioButtonToggled)

        self.downloadReportButton.clicked.connect(self.exportToExcel)

        self.runInference()

    def getAvailableCameras(self):
        available_cameras = []
        i = 0
        while True:
            cap = cv2.VideoCapture(i)
            if cap is None or not cap.isOpened():
                cap.release()
                break
            else:
                available_cameras.append(i)
                cap.release()
            i += 1
            
        available_cameras.append('rtsp://TAPOC200:TapoC200!@192.168.137.18:554/stream1')
        return available_cameras

    def getAvailableModels(self):
        # Get the directory of the current script
        directory = os.path.dirname(os.path.abspath(__file__))
        
        # Add the "models" subdirectory to the path
        models_directory = os.path.join(directory, 'models')
        
        # Return the full paths to the .pt files
        return [os.path.join(models_directory, f) for f in os.listdir(models_directory) if f.endswith('.pt')]

    def runInference(self):
        if hasattr(self, 'worker'):
            if self.worker.is_recording:
                # If currently recording, create a new video capture without stopping the recording
                self.worker.capture = cv2.VideoCapture(self.CameraSourceComboBox.currentData())

                # Update the worker's model index, confidence, and IOU threshold
                self.worker.model_index = self.model_index
                self.worker.conf = self.conf
                self.worker.iou = self.iou
            else:
                # If not recording, stop the current worker and release the camera
                self.worker.imageUpdated.disconnect()
                self.worker.errorOccurred.disconnect()
                self.thread.quit()
                self.video_capture.release()

                # Start a new video capture with the selected camera and create a new worker
                self.current_camera_index = self.CameraSourceComboBox.currentData()
                self.video_capture = cv2.VideoCapture(self.current_camera_index)

                self.thread = QThread(self)
                self.worker = VideoWorker(self.video_capture, self.model_index, self.conf, self.iou, self.models, self.maxID)
                self.worker.moveToThread(self.thread)

                self.worker.imageUpdated.connect(self.update_frame)
                self.worker.objectListUpdated.connect(self.update_list_widget)
                self.worker.errorOccurred.connect(self.handle_error)
                self.thread.started.connect(self.worker.run)

                self.thread.start()
        else:
            # If the worker doesn't exist, start a new video capture with the selected camera and create a new worker
            self.current_camera_index = self.CameraSourceComboBox.currentData()
            self.video_capture = cv2.VideoCapture(self.current_camera_index)

            self.thread = QThread(self)
            self.worker = VideoWorker(self.video_capture, self.model_index, self.conf, self.iou, self.models, self.maxID)
            self.worker.moveToThread(self.thread)

            self.worker.imageUpdated.connect(self.update_frame)
            self.worker.objectListUpdated.connect(self.update_list_widget)
            self.worker.errorOccurred.connect(self.handle_error)
            self.thread.started.connect(self.worker.run)

            self.thread.start()

    def update_frame(self, image):
        pixmap = QPixmap.fromImage(image)
        scaled_pixmap = pixmap.scaled(self.VideoWidget.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)  # Scale while keeping aspect ratio
        self.VideoWidget.setPixmap(scaled_pixmap)

    def update_list_widget(self, detected_objects):
        object_counts = {}
        for obj in detected_objects:
            object_name = obj[0]  # Extract the object name from the tuple
            object_counts[object_name] = object_counts.get(object_name, 0) + 1

        self.listWidget.clear()
        for obj, count in object_counts.items():
            self.listWidget.addItem(f"{obj} - {count}")

        # Calculate and display total objects
        total_objects = sum(object_counts.values())
        self.totalObjectsDetectedNumber.setText(str(total_objects))

        # Check alerts and display alertDialogBox if necessary
        with Database("selectAllAlerts") as database:
            alerts = database.selectAllAlerts()
            for item_name, alert_amount in alerts:
                if item_name == "All Classes":
                    if total_objects >= alert_amount:
                        self.alertDialogBox(item_name, alert_amount)
                else:
                    if object_counts.get(item_name, 0) >= alert_amount:
                        self.alertDialogBox(item_name, alert_amount)

    def onAdvancedSettingsClicked(self):
        advancedSettings = settingsDialog(self, self.models, self.model_index, self.conf, self.iou)
        advancedSettings.modelChanged.connect(self.updateModel)
        if advancedSettings.exec():
            self.conf = advancedSettings.conf
            self.iou = advancedSettings.iou
            self.model_index = advancedSettings.AIModelComboBox.currentIndex()
            self.runInference()

    def onSetAlertClicked(self):
        self.worker.open_alert_dialog()

    def updateModel(self, model_index):
        self.model_index = model_index

    def toggle_recording(self):
        if self.worker.is_recording:
            self.worker.stop_recording()
            self.StartRecordButton.setText("Start Recording")  # Update button text
        else:
            self.worker.start_recording()
            self.StartRecordButton.setText("Stop Recording")  # Update button text

    def maxIDFromDatabase(self):
        with Database("selectMaxID") as database:
            self.max_id = database.selectMaxID()
        return self.max_id

    def updateDateEdits(self):
        with Database("updateDateEdits") as database:

            # Retrieve the earliest and latest dates
            earliest_date_str, latest_date_str = database.updateDateEdits()

            # Ensure that there are records in the database
            if earliest_date_str and latest_date_str:    
                # Assuming dateTime format is "MM-DD-YYYY HH-MM-SS"
                earliest_date = QDate.fromString(earliest_date_str.split(" ")[0], "MM-dd-yyyy")
                latest_date = QDate.fromString(latest_date_str.split(" ")[0], "MM-dd-yyyy")

                # Set the DateEdits
                self.fromDateEdit.setDate(earliest_date)
                self.fromDateEdit.setMinimumDate(earliest_date)
                self.fromDateEdit.setMaximumDate(latest_date)

                self.toDateEdit.setDate(latest_date)
                self.toDateEdit.setMaximumDate(latest_date)
                self.toDateEdit.setMinimumDate(earliest_date)

    def updateToDateMinimum(self):
        # Set the minimum date of toDateEdit to the selected date of fromDateEdit
        self.toDateEdit.setMinimumDate(self.fromDateEdit.date())

    def radioButtonToggled(self):
        if self.dailyRadioButton.isChecked():
            self.toDateEdit.setVisible(True)
            self.dateEditMinusLabel.setVisible(True)
            self.label_3.setText("Average Daily Detected Litter")
        elif self.hourlyRadioButton.isChecked():
            self.toDateEdit.setVisible(False)
            self.dateEditMinusLabel.setVisible(False)
            self.label_3.setText("Average Hourly Detected Litter")

    def clickedFilterButton(self):
        with Database("clickedFilterButton") as database:
            from_date = self.fromDateEdit.date().toString("MM-dd-yyyy")
            to_date = self.toDateEdit.date().addDays(1).toString("MM-dd-yyyy") if self.dailyRadioButton.isChecked() else self.fromDateEdit.date().addDays(1).toString("MM-dd-yyyy")

            total_litter_count = database.count_rows_by_date_range(from_date, to_date)
            self.totalDetectedLitterLabel.setText(f"{total_litter_count}")

            if self.dailyRadioButton.isChecked():
                # Calculate days between dates
                from_date_obj = datetime.datetime.strptime(from_date, "%m-%d-%Y").date()
                to_date_obj = datetime.datetime.strptime(to_date, "%m-%d-%Y").date()
                num_days = (to_date_obj - from_date_obj).days

                if num_days > 0:  # Prevent division by zero
                    average_daily_litter_count = total_litter_count / num_days
                    self.averageDailyDetectedLitterLabel.setText(f"{average_daily_litter_count:.2f}")  # Format to 2 decimal places
                else:
                    self.averageDailyDetectedLitterLabel.setText("Invalid date range")
            elif self.hourlyRadioButton.isChecked():
                # Assume 24 hours in a day
                average_hourly_litter_count = total_litter_count / 24
                self.averageDailyDetectedLitterLabel.setText(f"{average_hourly_litter_count:.2f}")  # Format to 2 decimal places

            litter_data = database.litter_composition(from_date, to_date)
            self.totalRecyclableLitterLabel.setText(f"{litter_data['recyclable_percentage']:.2f}%")
            self.totalNonRecyclableLitterLabel.setText(f"{litter_data['non_recyclable_percentage']:.2f}%")

            recyclable_summary = database.litter_summary(from_date, to_date, True)
            self.recyclableLitterListWidget.clear()
            if recyclable_summary:
                for item in recyclable_summary:
                    display_text = f"{item['className']} - {item['count']}"
                    self.recyclableLitterListWidget.addItem(display_text)

            non_recyclable_summary = database.litter_summary(from_date, to_date, False)
            self.nonRecyclableLitterListWidget.clear()
            if non_recyclable_summary:
                for item in non_recyclable_summary:
                    display_text = f"{item['className']} - {item['count']}"
                    self.nonRecyclableLitterListWidget.addItem(display_text)

            litter_data_per_day = database.detected_litter_per_day(from_date, to_date, self.hourlyRadioButton.isChecked())
            if litter_data_per_day:
                self.createLineChart(litter_data_per_day)
            else:
                self.clearLineChart()

            count_per_class = database.count_class_names(from_date, to_date)
            if count_per_class:
                self.createBarChart(count_per_class)
            else:
                self.clearBarChart()

            percentage_per_class = database.get_class_percentages(from_date, to_date)
            if percentage_per_class:
                self.createPieChart(percentage_per_class)
            else:
                self.clearPieChart()

    def clearLineChart(self):
        for i in reversed(range(self.lineChartGridLayout.count())):
            self.lineChartGridLayout.itemAt(i).widget().setParent(None)

    def clearBarChart(self):
        for i in reversed(range(self.barChartGridLayout.count())):
            self.barChartGridLayout.itemAt(i).widget().setParent(None)

    def clearPieChart(self):
        for i in reversed(range(self.pieChartGridLayout.count())):
            self.pieChartGridLayout.itemAt(i).widget().setParent(None)

    def createLineChart(self, litter_data):
        # Process the data for the chart
        if self.dailyRadioButton.isChecked():
            x_values = [item['date'] for item in litter_data]
            x_label = 'Date'
            title = 'Detected Litter per Day'
        else:
            x_values = [item['hour'] for item in litter_data]
            x_label = 'Hour'
            title = 'Detected Litter per Hour'

        # Aggregate the counts for each hour
        counts = {}
        for item in litter_data:
            if self.hourlyRadioButton.isChecked():
                hour = item['hour']
            else:
                hour = item['date']
            count = int(item['count'])
            if hour not in counts:
                counts[hour] = 0
            counts[hour] += count

        y_values = list(counts.values())

        # Create the Matplotlib figure and axes
        self.lineChartFigure = Figure()
        ax = self.lineChartFigure.add_subplot(111)

        # Create the line chart
        ax.plot(x_values, y_values)

        # Customize the chart
        ax.set_xlabel(x_label)
        ax.set_ylabel('Detected Litter Count')
        ax.set_title(title)

        # Adjust font size based on figure size
        rcParams.update({'font.size': max(10, min(self.lineChartFigure.get_size_inches()))})

        # Adjust layout to fit the frame
        self.lineChartFigure.tight_layout()

        # Create a canvas widget to embed the chart
        canvas = FigureCanvas(self.lineChartFigure)

        # Clear previous content (if any) in the frame
        for i in reversed(range(self.lineChartGridLayout.count())):
            self.lineChartGridLayout.itemAt(i).widget().setParent(None)

        # Add the chart canvas to the frame's layout
        self.lineChartGridLayout.addWidget(canvas)  # Assuming you want to place it at the top
        canvas.draw()  # Draw the chart on the canvas

    def createBarChart(self, count_per_class):
        # Process the data for the chart
        class_names = [list(item.keys())[0] for item in count_per_class]
        counts = [int(list(item.values())[0]) for item in count_per_class]

        # Create the Matplotlib figure and axes
        self.barChartFigure = Figure()
        ax = self.barChartFigure.add_subplot(111)  

        # Create the bar chart
        ax.bar(class_names, counts)

        # Customize the chart
        ax.set_xlabel('Class Name')
        ax.set_ylabel('Count')
        ax.set_title('Count per Class')
        ax.set_xticklabels(class_names, rotation=45)

        # Adjust font size based on figure size
        rcParams.update({'font.size': max(10, min(self.barChartFigure.get_size_inches()))})

        # Adjust layout to fit the frame
        self.barChartFigure.tight_layout()

        # Create a canvas widget to embed the chart
        canvas = FigureCanvas(self.barChartFigure)

        # Clear previous content (if any) in the frame
        for i in reversed(range(self.barChartGridLayout.count())):
            self.barChartGridLayout.itemAt(i).widget().setParent(None)

        # Add the chart canvas to the frame's layout
        self.barChartGridLayout.addWidget(canvas)  # Assuming you want to place it at the top
        canvas.draw()  # Draw the chart on the canvas

    def createPieChart(self, percentage_per_class):
        # Process the data for the chart
        class_names = [list(item.keys())[0] for item in percentage_per_class]
        percentages = [float(list(item.values())[0].replace('%', '')) for item in percentage_per_class]

        # Create the Matplotlib figure and axes
        self.pieChartFigure = Figure()
        ax = self.pieChartFigure.add_subplot(111)  

        def label_function(val):
            return '' if val < 3 else f'{val:.1f}%'  # Omits labels for small percentages

        # Create the pie chart
        explode = [0.1]*len(class_names)  # Create a list of the same length as class_names, all values are 0.1
        wedges, texts, autotexts = ax.pie(percentages, autopct=label_function, explode=explode)

        # Create the legend labels with percentages
        legend_labels = [f'{name}: {percentage}%' for name, percentage in zip(class_names, percentages)]

        # Create the legend
        ax.legend(wedges, legend_labels, title="Class Names", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

        # Customize the chart
        ax.set_title('Percentage per Class')

        # Adjust font size based on figure size
        rcParams.update({'font.size': max(10, min(self.pieChartFigure.get_size_inches()))})

        # Adjust layout to fit the frame
        self.pieChartFigure.tight_layout()

        # Create a canvas widget to embed the chart
        canvas = FigureCanvas(self.pieChartFigure)

        # Clear previous content (if any) in the frame
        for i in reversed(range(self.pieChartGridLayout.count())):
            self.pieChartGridLayout.itemAt(i).widget().setParent(None)

        # Add the chart canvas to the frame's layout
        self.pieChartGridLayout.addWidget(canvas)  # Assuming you want to place it at the top
        canvas.draw()  # Draw the chart on the canvas

    def exportToExcel(self):
        with Database("exportToExcel") as database:
            from_date = self.fromDateEdit.date().toString("MM-dd-yyyy")
            to_date = self.toDateEdit.date().addDays(1).toString("MM-dd-yyyy") if self.dailyRadioButton.isChecked() else self.fromDateEdit.date().addDays(1).toString("MM-dd-yyyy")

            total_litter_count = database.count_rows_by_date_range(from_date, to_date)

            if total_litter_count == 0:
                QMessageBox.warning(self, "No Data to Export", "There is no data to export for the selected date range.")
                return

            if self.dailyRadioButton.isChecked():
                from_date_obj = datetime.datetime.strptime(from_date, "%m-%d-%Y").date()
                to_date_obj = datetime.datetime.strptime(to_date, "%m-%d-%Y").date()
                num_days = (to_date_obj - from_date_obj).days
                average_daily_litter_count = total_litter_count / num_days if num_days > 0 else 0
            else:
                average_hourly_litter_count = total_litter_count / 24

            litter_data = database.litter_composition(from_date, to_date)
            recyclable_summary = database.litter_summary(from_date, to_date, True)
            non_recyclable_summary = database.litter_summary(from_date, to_date, False)
            litter_data_per_day = database.detected_litter_per_day(from_date, to_date, self.hourlyRadioButton.isChecked())
            count_per_class = database.count_class_names(from_date, to_date)
            percentage_per_class = database.get_class_percentages(from_date, to_date)

            # Create a new workbook and select the active sheet
            workbook = Workbook()
            sheet = workbook.active

            # Write the data to the sheet
            bold_font = Font(bold=True)

            sheet.cell(row=1, column=1, value="From Date").font = bold_font
            sheet.cell(row=1, column=2, value=from_date)

            if self.dailyRadioButton.isChecked():
                sheet.cell(row=2, column=1, value="To Date").font = bold_font
                sheet.cell(row=2, column=2, value=to_date)

            sheet.cell(row=4, column=1, value="Total Detected Litter").font = bold_font
            sheet.cell(row=5, column=1, value=total_litter_count)

            if self.dailyRadioButton.isChecked():
                sheet.cell(row=6, column=1, value="Average Daily Detected Litter").font = bold_font
                sheet.cell(row=7, column=1, value=average_daily_litter_count)
            else:
                sheet.cell(row=6, column=1, value="Average Hourly Detected Litter").font = bold_font
                sheet.cell(row=7, column=1, value=average_hourly_litter_count)

            sheet.cell(row=8, column=1, value="Total Recyclable Litter (%)").font = bold_font
            sheet.cell(row=9, column=1, value=f"{litter_data['recyclable_percentage']:.2f}%")
            sheet.cell(row=10, column=1, value="Total Non-Recyclable Litter (%)").font = bold_font
            sheet.cell(row=11, column=1, value=f"{litter_data['non_recyclable_percentage']:.2f}%")

            sheet.cell(row=13, column=1, value="Recyclable Litter").font = bold_font
            for i, item in enumerate(recyclable_summary, start=14):
                sheet.cell(row=i, column=1, value=f"{item['className']} - {item['count']}")

            sheet.cell(row=13, column=2, value="Non-Recyclable Litter").font = bold_font
            for i, item in enumerate(non_recyclable_summary, start=14):
                sheet.cell(row=i, column=2, value=f"{item['className']} - {item['count']}")

            # Write the chart data to the sheet
            detected_litter_row = 1
            sheet.cell(row=detected_litter_row, column=7, value="Detected Litter per Day").font = bold_font
            for i, item in enumerate(litter_data_per_day, start=detected_litter_row + 1):
                sheet.cell(row=i, column=7, value=item['date'] if self.dailyRadioButton.isChecked() else item['hour'])
                sheet.cell(row=i, column=8, value=item['count'])

            count_per_class_row = detected_litter_row + len(litter_data_per_day) + 3
            sheet.cell(row=count_per_class_row, column=7, value="Count per Class").font = bold_font
            for i, item in enumerate(count_per_class, start=count_per_class_row + 1):
                class_name = list(item.keys())[0]
                count = list(item.values())[0]
                sheet.cell(row=i, column=7, value=class_name)
                sheet.cell(row=i, column=8, value=count)

            percentage_per_class_row = count_per_class_row + len(count_per_class) + 3
            sheet.cell(row=percentage_per_class_row, column=7, value="Percentage per Class").font = bold_font
            for i, item in enumerate(percentage_per_class, start=percentage_per_class_row + 1):
                class_name = list(item.keys())[0]
                percentage = list(item.values())[0]
                sheet.cell(row=i, column=7, value=class_name)
                sheet.cell(row=i, column=8, value=percentage)

            # Adjust column widths to fit the contents
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

            # Save the line chart as an image
            line_chart_path = "line_chart.png"
            if hasattr(self, 'lineChartFigure'):
                self.lineChartFigure.savefig(line_chart_path)
                line_chart_img = Image(line_chart_path)
                sheet.add_image(line_chart_img, "J2")

            # Save the bar chart as an image
            bar_chart_path = "bar_chart.png"
            if hasattr(self, 'barChartFigure'):
                self.barChartFigure.savefig(bar_chart_path)
                bar_chart_img = Image(bar_chart_path)
                sheet.add_image(bar_chart_img, "J" + str(count_per_class_row))

            # Save the pie chart as an image
            pie_chart_path = "pie_chart.png"
            if hasattr(self, 'pieChartFigure'):
                self.pieChartFigure.savefig(pie_chart_path)
                pie_chart_img = Image(pie_chart_path)
                sheet.add_image(pie_chart_img, "J" + str(percentage_per_class_row))

            # Create the "reports" folder if it doesn't exist
            reports_folder = "reports"
            if not os.path.exists(reports_folder):
                os.makedirs(reports_folder)

            # Generate a default file name based on the date range
            if self.dailyRadioButton.isChecked():
                default_file_name = f"litter_report_{from_date}_to_{to_date}.xlsx"
            else:
                default_file_name = f"litter_report_{from_date}.xlsx"
            default_file_path = os.path.join(reports_folder, default_file_name)

            # Save the workbook to a file
            file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", default_file_path, "Excel Files (*.xlsx)")
            if file_path:
                workbook.save(file_path)
                QMessageBox.information(self, "Export Successful", "Data exported to Excel successfully.")
            else:
                QMessageBox.warning(self, "Export Cancelled", "Export to Excel cancelled.")

            # Delete the temporary chart image files
            if os.path.exists(line_chart_path):
                os.remove(line_chart_path)
            if os.path.exists(bar_chart_path):
                os.remove(bar_chart_path)
            if os.path.exists(pie_chart_path):
                os.remove(pie_chart_path)
                   
    def alertDialogBox(self, className, alertAmount):
        if not self.alert_dialog_shown:
            self.alert_dialog_shown = True
            dlg = QMessageBox(self)
            dlg.setWindowTitle("Alert!")
            dlg.setText(f"{className} has reached the amount of {alertAmount}!")
            dlg.setStandardButtons(QMessageBox.Close)
            dlg.setIcon(QMessageBox.Warning)
            dlg.buttonClicked.connect(self.resetAlertDialogFlag)
            dlg.exec()

    def resetAlertDialogFlag(self):
        self.alert_dialog_shown = False

    def handle_error(self, message):
        print(f"Error occurred: {message}")

    def closeEvent(self, event):
        if hasattr(self, 'worker') and self.worker.is_recording:
            self.worker.stop_recording()  # Stop recording before closing
            self.thread.quit()  # Stop the thread
        super().closeEvent(event)

class settingsDialog(Ui_settingsDialog, QDialog):
    modelChanged = Signal(int)
    def __init__(self, parent=None, models=None, model_index=0, conf=0.40, iou=0.50):
        super().__init__(parent)
        self.setupUi(self)

        self.conf = conf
        self.iou = iou
        self.models = models if models is not None else []

        for model in self.models:
            self.AIModelComboBox.addItem(os.path.basename(model))

        self.AIModelComboBox.setCurrentIndex(model_index)

        # Set the range and initial value of the sliders
        self.ConfidenceThresholdHorizontalSlider.setRange(1, 100)
        self.ConfidenceThresholdHorizontalSlider.setValue(int(self.conf * 100))
        self.IoUThresholdHorizontalSlider.setRange(0, 95)
        self.IoUThresholdHorizontalSlider.setValue(int(self.iou * 100))

        # Set the initial text of the label to the value of conf and iou
        self.ConfidenceThresholdNumber.setText(str(self.conf))
        self.IoUThresholdNumber.setText(str(self.iou))

        # Connect the valueChanged signal to a slot
        self.ConfidenceThresholdHorizontalSlider.valueChanged.connect(self.updateConfidenceLabel)
        self.IoUThresholdHorizontalSlider.valueChanged.connect(self.updateIoULabel)

        # Connect the accepted signal to a slot
        self.buttonBox.accepted.connect(self.updateSliders)

    def updateConfidenceLabel(self, value):
        # Convert the value back to the range 0.01-1
        self.ConfidenceThresholdNumber.setText(str(value / 100))

    def updateIoULabel(self, value):
        # Convert the value back to the range 0-0.95
        self.IoUThresholdNumber.setText(str(value / 100))

    def updateSliders(self):
        # Update conf and iou to the current value of the sliders
        self.conf = self.ConfidenceThresholdHorizontalSlider.value() / 100
        self.iou = self.IoUThresholdHorizontalSlider.value() / 100

    def updateModel(self):
        self.modelChanged.emit(self.AIModelComboBox.currentIndex())

class setAlertDialog(Ui_setAlertDialog, QDialog):
    def __init__(self, parent=None, names=None):
        super().__init__(parent)
        self.setupUi(self)

        self.names = names if names is not None else {}

        # Populate the comboBox with the values from the names dictionary
        with Database("selectAllAlerts") as database:
            alerts = database.selectAllAlerts()
            existing_items = set(item_name for item_name, _ in alerts)

            self.comboBox.addItem("All Classes")  # adds the 'All Classes' option

            # Add items from the names dictionary that are not in the database
            for name in self.names.values():
                if name not in existing_items:
                    self.comboBox.addItem(name)

            # Sort the items in the comboBox alphabetically
            self.comboBox.model().sort(0)

        self.populate_rows_from_database()
        
    def save_to_database(self, item_name, amount):
        with Database("insertAlert") as database:
            database.insertAlert(item_name, int(amount))

    def remove_from_database(self, item_name):
        with Database("removeAlert") as database:
            database.removeAlert(item_name)

    def populate_rows_from_database(self):
        with Database("selectAllAlerts") as database:
            alerts = database.selectAllAlerts()
            for item_name, amount in alerts:
                self.add_row(item_name, amount)

app = QtWidgets.QApplication(sys.argv)

window = MainWindow()
window.show()
app.exec()